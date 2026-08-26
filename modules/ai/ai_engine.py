import os
import io
from pathlib import Path
import google.generativeai as genai
from pypdf import PdfReader
from docx import Document
import openpyxl
from PIL import Image
from core.config import GEMINI_API_KEY, logger
from core.database import save_ai_message, get_ai_history, clear_ai_history

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

GEMINI_SYSTEM_INSTRUCTION = """
Siz Google tomonidan yaratilgan eng mukammal, aqlli va xushmuomala sun'iy intellektsiz (Gemini AI).
Foydalanuvchining barcha savollariga aniq, mantiqiy, to'liq va tushunarli qilib javob bering.

Javob berish qoidalari:
1. Til: Foydalanuvchi qaysi tilda murojaat qilsa (O'zbek, Rus, Ingliz va boshqalar), o'sha tilda sof va adabiy javob bering.
2. Formatlash: Telegram uchun mos Markdown formatidan foydalaning (Sarlavhalar, **qalin**, `kod bloklari`, jadvallar, ro'yxatlar).
3. Dasturlash: Kod yozganda eng yaxshi amaliyotlar (clean code), tushuntirishlar va to'liq ishchi yechimlarni taqdim eting.
4. Tahlil: Rasm, hujjat (PDF, Word, Excel, kod) kelganda uning barcha nozik jihatlarini chuqur tahlil qiling.
5. Samimiy va professional uslubda muloqot qiling.
"""

def get_gemini_model(model_name: str = "gemini-3.6-flash"):
    generation_config = {
        "temperature": 0.7,
        "top_p": 0.95,
        "top_k": 40,
        "max_output_tokens": 8192,
    }
    return genai.GenerativeModel(
        model_name=model_name,
        system_instruction=GEMINI_SYSTEM_INSTRUCTION,
        generation_config=generation_config
    )

async def ask_gemini_chat(user_id: int, prompt: str) -> str:
    try:
        if not GEMINI_API_KEY:
            return "⚠️ Gemini API kaliti topilmadi."

        history_rows = get_ai_history(user_id, limit=12)
        formatted_history = []
        for r in history_rows:
            formatted_history.append({
                "role": "user" if r["role"] == "user" else "model",
                "parts": [r["content"]]
            })

        try:
            model = get_gemini_model("gemini-3.6-flash")
            chat = model.start_chat(history=formatted_history)
            response = chat.send_message(prompt)
        except Exception:
            model = get_gemini_model("gemini-3.7-flash")
            chat = model.start_chat(history=formatted_history)
            response = chat.send_message(prompt)
        
        reply_text = response.text
        save_ai_message(user_id, "user", prompt)
        save_ai_message(user_id, "model", reply_text)
        return reply_text
    except Exception as e:
        logger.error(f"Gemini Chat error: {e}")
        return f"⚠️ AI xatolik: {str(e)}"

async def analyze_image_with_ai(user_id: int, image_path: str, prompt: str = "") -> str:
    try:
        if not GEMINI_API_KEY:
            return "⚠️ Gemini API kaliti topilmadi."
            
        user_prompt = prompt.strip() if prompt.strip() else "Ushbu rasmni to'liq tahlil qilib, undagi ma'lumotlar, matnlar yoki savollarga batafsil javob bering:"
        img = Image.open(image_path)
        
        try:
            model = get_gemini_model("gemini-3.6-flash")
            response = model.generate_content([user_prompt, img])
        except Exception:
            model = get_gemini_model("gemini-3.7-flash")
            response = model.generate_content([user_prompt, img])
            
        reply_text = response.text
        save_ai_message(user_id, "user", f"[Rasm yuborildi]: {user_prompt}")
        save_ai_message(user_id, "model", reply_text)
        return reply_text
    except Exception as e:
        logger.error(f"Gemini Vision error: {e}")
        return f"⚠️ Rasm tahlilida xatolik: {str(e)}"

async def process_voice_with_ai(user_id: int, audio_path: str) -> str:
    try:
        if not GEMINI_API_KEY:
            return "⚠️ Gemini API kaliti topilmadi."
            
        model = get_gemini_model("gemini-3.6-flash")
        audio_file = genai.upload_file(path=audio_path)
        prompt = "Ushbu ovozli xabarni eshiting va foydalanuvchining gapiga to'liq, mazmunli va foydali javob qaytaring."
        response = model.generate_content([prompt, audio_file])
        
        reply_text = response.text
        save_ai_message(user_id, "user", "[Ovozli xabar yuborildi]")
        save_ai_message(user_id, "model", reply_text)
        return reply_text
    except Exception as e:
        logger.error(f"Gemini Voice error: {e}")
        return f"⚠️ Ovoz tahlilida xatolik: {str(e)}"

async def analyze_document_with_ai(user_id: int, file_path: str, user_question: str = "") -> str:
    ext = Path(file_path).suffix.lower()
    extracted_text = ""
    
    try:
        if ext == ".pdf":
            reader = PdfReader(file_path)
            for i, page in enumerate(reader.pages[:30]):
                t = page.extract_text()
                if t:
                    extracted_text += f"\n--- Sahifa {i+1} ---\n" + t
                    
        elif ext in [".docx", ".doc"]:
            doc = Document(file_path)
            extracted_text = "\n".join([p.text for p in doc.paragraphs if p.text])
            
        elif ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet in wb.sheetnames[:5]:
                ws = wb[sheet]
                extracted_text += f"\n--- Varaq: {sheet} ---\n"
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(v) for v in row if v is not None]
                    if row_vals:
                        extracted_text += "\t".join(row_vals) + "\n"
                        
        elif ext in [".txt", ".py", ".js", ".html", ".css", ".json", ".csv", ".md", ".cpp", ".java", ".sql"]:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                extracted_text = f.read()
                
        if not extracted_text.strip():
            return "⚠️ Hujjatdan matn ajratib olinmadi yoki fayl bo'sh."
            
        q = user_question.strip() if user_question.strip() else "Ushbu hujjatning to'liq tahlilini, xulosasini va muhim nuqtalarini yozib bering:"
        prompt = f"HUJJAT MATNI:\n{extracted_text[:25000]}\n\nFOYDALANUVCHI SAVOLI:\n{q}"
        
        return await ask_gemini_chat(user_id, prompt)
    except Exception as e:
        logger.error(f"Doc Analysis error: {e}")
        return f"⚠️ Hujjat tahlilida xatolik: {str(e)}"
